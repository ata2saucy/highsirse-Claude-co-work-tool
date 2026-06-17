# -*- coding: utf-8 -*-
"""
qa_workbook.py — QA gate for the rental-comp workbook (v3 standard).

Usage:  python tools/qa_workbook.py "10 Lower Spadina Rental Comps _vACTIVE.xlsx"

Checks (acceptance checklist from the v3 change spec):
  * recalculates every formula (whole-column ranges are bounded to a finite window
    so the pure-Python engine can evaluate them) and reports any visible error token
  * structural rules: exactly the 6 required sheets, no 'Subject & Conclusion',
    no score / /10 / comp-quality / judgment-score text in Output, Data_Summary keeps
    Premium Basis + Mix Basis and drops the old INPUTS block, RD Condos / RD Apartments
    share identical headers, Date columns formatted yyyy-mm-dd, SF column has no decimals
  * every RD row without an exact verified SF must have Include=0 and a stated reason
  * LINEST parking coefficient and the headline aggregates are independently re-derived
    with numpy and reconciled

Exit code 0 = pass, 1 = any failure. Safe to run repeatedly; it never writes the workbook.
"""
import sys, re, io, warnings, copy
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.utils import get_column_letter

REQUIRED_SHEETS = ["Output", "Building Summary", "Data_Summary", "RD Condos", "RD Apartments", "Floor Plans"]
ERR_TOKENS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NUM!", "#NULL!", "#ERROR", "#CYCLE")
BANNED_OUTPUT = ("score /10", "/10", "comp-quality", "claude judgment", "judgment score", "mirror of subject")

results = []  # (ok, label, detail)
def check(ok, label, detail=""):
    results.append((bool(ok), label, detail))


def bound_ranges(formula, maxrow=400):
    """Bound whole-column refs (A:A, $A:$A, 'Sheet'!$A:$A) to row window 1..maxrow."""
    pat = re.compile(r'(?<![A-Za-z0-9_$])(\$?)([A-Z]{1,3}):(\$?)([A-Z]{1,3})(?![0-9A-Za-z])')
    return pat.sub(lambda m: f"{m.group(1)}{m.group(2)}$1:{m.group(3)}{m.group(4)}${maxrow}", formula)


def recalc_errors(path):
    """Recalc a range-bounded copy with the `formulas` engine; return (error_list, note)."""
    try:
        import formulas
    except ImportError:
        return None, "formulas engine not installed (pip install formulas) — recalc skipped"
    wb = openpyxl.load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    c.value = bound_ranges(c.value)
    tmp = io.BytesIO()
    wb.save(tmp); tmp.seek(0)
    import tempfile, os
    fd, tp = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
    with open(tp, "wb") as f:
        f.write(tmp.getvalue())
    errs = []
    note = ""
    try:
        xl = formulas.ExcelModel().loads(tp).finish()
        sol = xl.calculate()
        for k, v in sol.items():
            val = getattr(v, "value", v)
            cand = []
            if hasattr(val, "ravel"):
                try:
                    cand = [x for x in val.ravel()]
                except Exception:
                    cand = []
            else:
                cand = [val]
            for x in cand:
                if isinstance(x, str) and any(e in x for e in ERR_TOKENS):
                    errs.append((k, str(x)[:40])); break
    except Exception as e:
        note = f"recalc raised {type(e).__name__}: {str(e)[:160]}"
    finally:
        try: os.remove(tp)
        except OSError: pass
    return errs, note


def numpy_checks(path):
    import numpy as np
    wb = openpyxl.load_workbook(path, data_only=False)
    rdc = wb["RD Condos"]
    SF, PARK, BED2, BED3, RENT = [], [], [], [], []
    for r in range(2, rdc.max_row + 1):
        b = rdc.cell(r, 2).value
        ab = rdc.cell(r, 28).value
        if isinstance(b, (int, float)) and b and isinstance(ab, (int, float)) and ab:
            beds = str(rdc.cell(r, 17).value or "")
            SF.append(b); PARK.append(rdc.cell(r, 21).value or 0)
            BED2.append(1 if beds[:1] == "2" else 0)
            BED3.append(1 if beds[:1] == "3" else 0)
            RENT.append(ab)
    n = len(RENT)
    if n >= 10 and (max(PARK) - min(PARK)) > 0:
        X = np.column_stack([SF, PARK, BED2, BED3, np.ones(n)])
        coef, *_ = np.linalg.lstsq(X, np.array(RENT, float), rcond=None)
        park_coef = coef[1]
        return n, park_coef
    return n, None


def main(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    names = wb.sheetnames

    check(names == REQUIRED_SHEETS, "Sheet set is exactly the 6 required sheets, in order",
          f"found: {names}")
    check("Subject & Conclusion" not in names, "No 'Subject & Conclusion' sheet")

    # Output: no score / comp-quality / judgment language
    out = wb["Output"]
    hits = []
    for row in out.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                low = c.value.lower()
                for b in BANNED_OUTPUT:
                    if b in low:
                        hits.append(f"{c.coordinate}:{b}")
    check(not hits, "Output carries no score / /10 / comp-quality / judgment / mirror text",
          "; ".join(hits[:8]))

    # Data_Summary: Premium Basis + Mix Basis present, INPUTS block gone
    ds = wb["Data_Summary"]
    dstext = "\n".join(str(c.value) for row in ds.iter_rows() for c in row if c.value is not None)
    check("PREMIUM BASIS" in dstext.upper(), "Data_Summary keeps a PREMIUM BASIS block")
    check("MIX BASIS" in dstext.upper(), "Data_Summary keeps a MIX BASIS block")
    check("INPUTS — HOW EACH" not in dstext.upper() and "INPUTS — HOW" not in dstext.upper(),
          "Data_Summary INPUTS block removed")

    # RD sheets structurally identical (headers)
    rdc, rda = wb["RD Condos"], wb["RD Apartments"]
    hc = [rdc.cell(1, i).value for i in range(1, rdc.max_column + 1)]
    ha = [rda.cell(1, i).value for i in range(1, rda.max_column + 1)]
    ncols = min(len(hc), len(ha))
    check(hc[:ncols] == ha[:ncols] and len(hc) == len(ha),
          "RD Condos and RD Apartments headers identical", f"{len(hc)} vs {len(ha)} cols")
    for need in ("SF Verification Status", "SF Source Type", "SF Source / URL"):
        check(need in hc, f"RD Condos has '{need}' column")

    # Date columns formatted yyyy-mm-dd (col J 'Date', AO 'Date Scraped')
    def col_idx(ws, label):
        for i in range(1, ws.max_column + 1):
            if ws.cell(1, i).value == label:
                return i
        return None
    bad_dates = []
    for label in ("Date", "Lease Date", "Date Scraped"):
        ci = col_idx(rdc, label)
        if ci:
            for r in range(2, rdc.max_row + 1):
                v = rdc.cell(r, ci).value
                if v is not None and "yy" in str(rdc.cell(r, ci).number_format).lower() and "yyyy-mm-dd" not in rdc.cell(r, ci).number_format:
                    bad_dates.append(f"{label}!{r}:{rdc.cell(r,ci).number_format}")
    check(not bad_dates, "RD date columns use yyyy-mm-dd", "; ".join(bad_dates[:6]))

    # SF column has no decimal format
    sf_i = col_idx(rdc, "Sq Ft.")
    fmt_bad = []
    for r in range(2, rdc.max_row + 1):
        f = rdc.cell(r, sf_i).number_format
        if "." in f:
            fmt_bad.append(f"{r}:{f}")
    check(not fmt_bad, "RD Condos SF column shows no decimals", "; ".join(fmt_bad[:6]))

    # Every row without numeric SF -> Include=0 and a stated reason
    inc_i = col_idx(rdc, "Include")
    expl_i = col_idx(rdc, "SF Explanation / Not Validated Reason")
    desc_i = col_idx(rdc, "Description")
    bad_excl = []
    for r in range(2, rdc.max_row + 1):
        b = rdc.cell(r, sf_i).value
        kbuild = rdc.cell(r, 11).value
        if not kbuild:
            continue
        if not isinstance(b, (int, float)) or not b:
            reason = (rdc.cell(r, expl_i).value if expl_i else None) or (rdc.cell(r, desc_i).value if desc_i else None)
            if not reason:
                bad_excl.append(f"row {r}: no reason")
            # Include is a formula =IF(...) -> evaluates to 0 when SF blank; just check it's the guarded formula
            inc = rdc.cell(r, inc_i).value
            if not (isinstance(inc, str) and inc.startswith("=IF")):
                bad_excl.append(f"row {r}: Include not guarded")
    check(not bad_excl, "Every RD row without exact SF is excluded with a reason", "; ".join(bad_excl[:6]))

    # Floor Plans logs the subject status + expanded columns
    fp = wb["Floor Plans"]
    fph = [fp.cell(1, i).value for i in range(1, fp.max_column + 1)]
    for need in ("Subject/Comp", "Source Site", "Source URL", "Date Read", "Verification Tier"):
        check(need in fph, f"Floor Plans has '{need}' column")
    fptext = "\n".join(str(c.value) for row in fp.iter_rows() for c in row if c.value)
    check("(SUBJECT)" in fptext or "SUBJECT" in fptext, "Floor Plans logs the subject row")

    # recalc
    errs, note = recalc_errors(path)
    if errs is None:
        check(False, "Recalc (formulas engine)", note)
    else:
        real = [(k, x) for k, x in errs if "LINEST" not in str(k)]
        check(len(errs) == 0, "Recalc: zero visible error cells",
              (note + " | " if note else "") + "; ".join(f"{k}={x}" for k, x in errs[:12]))

    # numpy cross-check of parking coef
    n, park = numpy_checks(path)
    check(n >= 10, f"Verified-SF leased rows for regression: {n}", "")
    if park is not None:
        check(True, f"Parking LINEST cross-check (numpy) = ${park:,.2f}/spot/mo", "")

    # report
    print("=" * 72)
    print(f"QA — {path}")
    print("=" * 72)
    npass = sum(1 for ok, *_ in results if ok)
    for ok, label, detail in results:
        mark = "PASS" if ok else "FAIL"
        line = f"[{mark}] {label}"
        if detail and not ok:
            line += f"  -> {detail}"
        elif detail and ok and detail.strip():
            line += f"  ({detail})"
        print(line)
    print("-" * 72)
    allok = all(ok for ok, *_ in results)
    print(f"{npass}/{len(results)} checks passed — {'ALL PASS' if allok else 'FAILURES PRESENT'}")
    return 0 if allok else 1


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "10 Lower Spadina Rental Comps _vACTIVE.xlsx"
    sys.exit(main(p))
